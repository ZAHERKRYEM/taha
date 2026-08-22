"""
بناء ملف Excel واحد يحوي جميع بيانات النظام (حلقات، طلاب، أساتذة، دورات،
وكل سجلات الحضور بأنواعها) في شيتات منفصلة، لإرساله كنسخة فورية لتلغرام
عند الطلب (بخلاف النسخة الاحتياطية اليومية لقاعدة البيانات كاملة).
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    Teacher, Circle, Student, Attendance, TeacherAttendance,
    Course, CourseAttendance, CourseTeacherAttendance, TeacherDailyAttendance,
)

HEADER_FILL = PatternFill(start_color='1F6F54', end_color='1F6F54', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def _write_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title=title[:31])  # حد أقصى 31 حرف لاسم الشيت في Excel
    ws.sheet_view.rightToLeft = True
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(row)
    # عرض أعمدة تلقائي تقريبي
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            val = row[col_idx - 1] if col_idx - 1 < len(row) else ''
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)
    ws.freeze_panes = 'A2'
    return ws


def build_full_export_workbook():
    """يبني ويعيد Workbook (openpyxl) يحوي جميع بيانات النظام في شيتات منفصلة."""
    wb = Workbook()
    wb.remove(wb.active)  # إزالة الشيت الافتراضي الفارغ

    # ---- الحلقات ----
    circles = Circle.objects.select_related('teacher').order_by('display_order', 'name')
    _write_sheet(
        wb, 'الحلقات',
        ['المعرّف', 'اسم الحلقة', 'الأستاذ المسؤول', 'الوصف', 'عدد الطلاب', 'تاريخ الإنشاء'],
        [
            [c.id, c.name, c.teacher.name if c.teacher else '', c.description,
             c.student_count, c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else '']
            for c in circles
        ],
    )

    # ---- الطلاب ----
    students = Student.objects.select_related('circle').order_by('circle__name', 'name')
    _write_sheet(
        wb, 'الطلاب',
        ['المعرّف', 'اسم الطالب', 'الحلقة', 'رقم الهاتف', 'ملاحظات', 'تاريخ الانضمام'],
        [
            [s.id, s.name, s.circle.name if s.circle else '', s.phone, s.notes,
             s.joined_at.strftime('%Y-%m-%d %H:%M') if s.joined_at else '']
            for s in students
        ],
    )

    # ---- الأساتذة ----
    teachers = Teacher.objects.prefetch_related('circles', 'courses').order_by('name')
    _write_sheet(
        wb, 'الأساتذة',
        ['المعرّف', 'الاسم', 'رقم الهاتف', 'الحلقات المرتبطة', 'الدورات المرتبطة'],
        [
            [t.id, t.name, t.phone,
             ', '.join(c.name for c in t.circles.all()),
             ', '.join(c.name for c in t.courses.all())]
            for t in teachers
        ],
    )

    # ---- الدورات والدروس الإضافية ----
    courses = Course.objects.prefetch_related('teachers', 'students').order_by('name')
    _write_sheet(
        wb, 'الدورات والدروس',
        ['المعرّف', 'الاسم', 'الأساتذة', 'الوصف', 'عدد الطلاب', 'تاريخ الإنشاء'],
        [
            [c.id, c.name, ', '.join(t.name for t in c.teachers.all()), c.description,
             c.student_count, c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else '']
            for c in courses
        ],
    )

    # ---- حضور الطلاب في الحلقات ----
    att = Attendance.objects.select_related('student', 'circle').order_by('-date', 'circle__name', 'student__name')
    _write_sheet(
        wb, 'حضور الحلقات',
        ['التاريخ', 'الحلقة', 'اسم الطالب', 'الحالة', 'وقت التسجيل'],
        [
            [a.date.isoformat(), a.circle.name if a.circle else '', a.student.name if a.student else '',
             a.get_status_display(), a.recorded_at.strftime('%Y-%m-%d %H:%M') if a.recorded_at else '']
            for a in att
        ],
    )

    # ---- حضور الأساتذة في الحلقات ----
    tatt = TeacherAttendance.objects.select_related('teacher', 'circle').order_by('-date', 'circle__name')
    _write_sheet(
        wb, 'حضور أساتذة الحلقات',
        ['التاريخ', 'الحلقة', 'اسم الأستاذ', 'الحالة'],
        [
            [t.date.isoformat(), t.circle.name if t.circle else '', t.teacher.name if t.teacher else '',
             t.get_status_display()]
            for t in tatt
        ],
    )

    # ---- حضور الطلاب في الدورات ----
    catt = CourseAttendance.objects.select_related('student', 'course').order_by('-date', 'course__name', 'student__name')
    _write_sheet(
        wb, 'حضور الدورات',
        ['التاريخ', 'الدورة/الدرس', 'اسم الطالب', 'الحالة'],
        [
            [c.date.isoformat(), c.course.name if c.course else '', c.student.name if c.student else '',
             c.get_status_display()]
            for c in catt
        ],
    )

    # ---- حضور الأساتذة في الدورات ----
    ctatt = CourseTeacherAttendance.objects.select_related('teacher', 'course').order_by('-date', 'course__name')
    _write_sheet(
        wb, 'حضور أساتذة الدورات',
        ['التاريخ', 'الدورة/الدرس', 'اسم الأستاذ', 'الحالة'],
        [
            [c.date.isoformat(), c.course.name if c.course else '', c.teacher.name if c.teacher else '',
             c.get_status_display()]
            for c in ctatt
        ],
    )

    # ---- التفقد العام للأساتذة غير المرتبطين ----
    gen = TeacherDailyAttendance.objects.select_related('teacher').order_by('-date')
    _write_sheet(
        wb, 'التفقد العام للأساتذة',
        ['التاريخ', 'اسم الأستاذ', 'الحالة'],
        [
            [g.date.isoformat(), g.teacher.name if g.teacher else '', g.get_status_display()]
            for g in gen
        ],
    )

    return wb


def build_full_export_bytes():
    """يبني الملف ويعيده كـ bytes جاهزة للإرسال أو الحفظ، دون الحاجة لملف مؤقت على القرص."""
    wb = build_full_export_workbook()
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
