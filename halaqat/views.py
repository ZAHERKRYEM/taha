import csv
import datetime
import io
from random import randint
import threading
import time

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.db import transaction
from django.views.decorators.http import require_POST

from .models import (
    Teacher, Circle, Student, Attendance, TeacherAttendance,
    Course, CourseAttendance, CourseTeacherAttendance, TeacherDailyAttendance,
)
from .forms import CircleForm, StudentForm, CourseForm, WhatsAppAbsentForm
from .telegram_utils import send_telegram_message, send_telegram_document_bytes
from .whatsapp_utils import normalize_chat_id, send_whatsapp_text


@login_required
def circles_list(request):
    circles = Circle.objects.select_related('teacher').order_by('display_order', 'name', 'id')
    today = timezone.now().date()

    circle_cards = []
    for circle in circles:
        total_students = circle.student_count
        today_records = Attendance.objects.filter(circle=circle, date=today)
        present_today = today_records.filter(status__in=['present', 'late']).count()
        rate = None
        if today_records.exists():
            counted = today_records.exclude(status='excused').count()
            rate = round((present_today / counted) * 100) if counted else None

        progress = circle.attendance_progress(today)
        if progress['total'] == 0:
            card_status = 'empty'
        elif progress['complete']:
            card_status = 'complete'
        else:
            card_status = 'incomplete'

        circle_cards.append({
            'circle': circle,
            'student_count': total_students,
            'today_rate': rate,
            'card_status': card_status,
            'progress_total': progress['total'],
            'progress_recorded': progress['recorded'],
        })

    context = {
        'circle_cards': circle_cards,
        'total_circles': circles.count(),
        'total_students': Student.objects.count(),
        'today': today,
        'active_nav': 'circles',
    }
    return render(request, 'halaqat/circles_list.html', context)


@login_required
def circle_detail(request, circle_id):
    circle = get_object_or_404(Circle, pk=circle_id)
    students = circle.students.all()

    # تحديد التاريخ المعروض (اليوم افتراضياً، أو من ?date=YYYY-MM-DD)
    date_str = request.GET.get('date')
    if date_str:
        try:
            selected_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()

    # جلب الحالات الحالية لهذا اليوم
    existing = {
        a.student_id: a.status
        for a in Attendance.objects.filter(circle=circle, date=selected_date)
    }

    student_rows = []
    for student in students:
        student_rows.append({
            'student': student,
            'status': existing.get(student.id, ''),  # فارغ = لم يُسجَّل بعد
        })

    present_count = sum(1 for s in student_rows if s['status'] == 'present')
    late_count = sum(1 for s in student_rows if s['status'] == 'late')
    absent_count = sum(1 for s in student_rows if s['status'] == 'absent')
    excused_count = sum(1 for s in student_rows if s['status'] == 'excused')

    teacher_status = ''
    if circle.teacher:
        ta = TeacherAttendance.objects.filter(
            teacher=circle.teacher, circle=circle, date=selected_date
        ).first()
        teacher_status = ta.status if ta else ''

    # للتنقل بين الحلقات دون الخروج من الصفحة (أزرار سابق/تالي + قائمة منسدلة)
    all_circles = list(Circle.objects.order_by('display_order', 'name', 'id'))
    ids = [c.id for c in all_circles]
    idx = ids.index(circle.id) if circle.id in ids else None
    prev_circle = all_circles[idx - 1] if idx is not None and idx > 0 else None
    next_circle = all_circles[idx + 1] if idx is not None and idx < len(all_circles) - 1 else None

    context = {
        'circle': circle,
        'student_rows': student_rows,
        'selected_date': selected_date,
        'present_count': present_count,
        'late_count': late_count,
        'absent_count': absent_count,
        'excused_count': excused_count,
        'attendance_rate_30d': circle.attendance_rate(30),
        'teacher_status': teacher_status,
        'all_circles': all_circles,
        'prev_circle': prev_circle,
        'next_circle': next_circle,
        'active_nav': 'circles',
    }
    return render(request, 'halaqat/circle_detail.html', context)


@login_required
@require_POST
def save_attendance(request, circle_id):
    """حفظ فوري لحالة حضور طالب أو الأستاذ عبر AJAX - بدون زر حفظ."""
    circle = get_object_or_404(Circle, pk=circle_id)
    kind = request.POST.get('kind', 'student')
    entity_id = request.POST.get('entity_id')
    status = request.POST.get('status')
    date_str = request.POST.get('date')

    if status not in dict(Attendance.STATUS_CHOICES):
        return JsonResponse({'ok': False, 'error': 'حالة غير صالحة'}, status=400)

    try:
        selected_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return JsonResponse({'ok': False, 'error': 'تاريخ غير صالح'}, status=400)

    if kind == 'teacher':
        if not circle.teacher or str(circle.teacher.id) != str(entity_id):
            return JsonResponse({'ok': False, 'error': 'لا يوجد أستاذ مطابق لهذه الحلقة'}, status=400)
        TeacherAttendance.objects.update_or_create(
            teacher=circle.teacher, circle=circle, date=selected_date,
            defaults={'status': status},
        )
    else:
        student = get_object_or_404(Student, pk=entity_id, circle=circle)
        Attendance.objects.update_or_create(
            student=student, date=selected_date,
            defaults={'circle': circle, 'status': status},
        )

    records = Attendance.objects.filter(circle=circle, date=selected_date)
    counts = {
        'present': records.filter(status='present').count(),
        'late': records.filter(status='late').count(),
        'absent': records.filter(status='absent').count(),
        'excused': records.filter(status='excused').count(),
    }
    return JsonResponse({'ok': True, 'counts': counts})


# ============================================================
#  الدورات والدروس الإضافية (فئة محددة من الطلاب، حضور مستقل)
# ============================================================

@login_required
def courses_list(request):
    courses = Course.objects.prefetch_related('teachers', 'students')
    today = timezone.now().date()

    course_cards = []
    for course in courses:
        total_students = course.student_count
        today_records = CourseAttendance.objects.filter(course=course, date=today)
        present_today = today_records.filter(status__in=['present', 'late']).count()
        rate = None
        if today_records.exists():
            counted = today_records.exclude(status='excused').count()
            rate = round((present_today / counted) * 100) if counted else None

        progress = course.attendance_progress(today)
        if progress['total'] == 0:
            card_status = 'empty'
        elif progress['complete']:
            card_status = 'complete'
        else:
            card_status = 'incomplete'

        course_cards.append({
            'course': course,
            'student_count': total_students,
            'today_rate': rate,
            'card_status': card_status,
        })

    context = {
        'course_cards': course_cards,
        'total_courses': courses.count(),
        'today': today,
        'active_nav': 'courses',
    }
    return render(request, 'halaqat/courses_list.html', context)


@login_required
def course_detail(request, course_id):
    course = get_object_or_404(Course, pk=course_id)
    students = course.students.select_related('circle').order_by('name')

    date_str = request.GET.get('date')
    if date_str:
        try:
            selected_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()

    existing = {
        a.student_id: a.status
        for a in CourseAttendance.objects.filter(course=course, date=selected_date)
    }

    student_rows = []
    for student in students:
        student_rows.append({
            'student': student,
            'status': existing.get(student.id, ''),
        })

    present_count = sum(1 for s in student_rows if s['status'] == 'present')
    late_count = sum(1 for s in student_rows if s['status'] == 'late')
    absent_count = sum(1 for s in student_rows if s['status'] == 'absent')
    excused_count = sum(1 for s in student_rows if s['status'] == 'excused')

    teacher_rows = []
    for teacher in course.teachers.all():
        cta = CourseTeacherAttendance.objects.filter(
            teacher=teacher, course=course, date=selected_date
        ).first()
        teacher_rows.append({
            'teacher': teacher,
            'status': cta.status if cta else '',
        })

    all_courses = list(Course.objects.order_by('name'))
    ids = [c.id for c in all_courses]
    idx = ids.index(course.id) if course.id in ids else None
    prev_course = all_courses[idx - 1] if idx is not None and idx > 0 else None
    next_course = all_courses[idx + 1] if idx is not None and idx < len(all_courses) - 1 else None

    context = {
        'course': course,
        'student_rows': student_rows,
        'selected_date': selected_date,
        'present_count': present_count,
        'late_count': late_count,
        'absent_count': absent_count,
        'excused_count': excused_count,
        'attendance_rate_30d': course.attendance_rate(30),
        'teacher_rows': teacher_rows,
        'all_courses': all_courses,
        'prev_course': prev_course,
        'next_course': next_course,
        'active_nav': 'courses',
    }
    return render(request, 'halaqat/course_detail.html', context)


@login_required
@require_POST
def save_course_attendance(request, course_id):
    """حفظ فوري لحالة حضور طالب أو الأستاذ ضمن دورة/درس إضافي عبر AJAX."""
    course = get_object_or_404(Course, pk=course_id)
    kind = request.POST.get('kind', 'student')
    entity_id = request.POST.get('entity_id')
    status = request.POST.get('status')
    date_str = request.POST.get('date')

    if status not in dict(Attendance.STATUS_CHOICES):
        return JsonResponse({'ok': False, 'error': 'حالة غير صالحة'}, status=400)

    try:
        selected_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return JsonResponse({'ok': False, 'error': 'تاريخ غير صالح'}, status=400)

    if kind == 'teacher':
        teacher = get_object_or_404(Teacher, pk=entity_id)
        if not course.teachers.filter(pk=teacher.pk).exists():
            return JsonResponse({'ok': False, 'error': 'لا يوجد أستاذ مطابق لهذه الدورة'}, status=400)
        CourseTeacherAttendance.objects.update_or_create(
            teacher=teacher, course=course, date=selected_date,
            defaults={'status': status},
        )
    else:
        student = get_object_or_404(Student, pk=entity_id)
        if not course.students.filter(pk=student.pk).exists():
            return JsonResponse({'ok': False, 'error': 'الطالب ليس ضمن هذه الدورة'}, status=400)
        CourseAttendance.objects.update_or_create(
            student=student, course=course, date=selected_date,
            defaults={'status': status},
        )

    records = CourseAttendance.objects.filter(course=course, date=selected_date)
    counts = {
        'present': records.filter(status='present').count(),
        'late': records.filter(status='late').count(),
        'absent': records.filter(status='absent').count(),
        'excused': records.filter(status='excused').count(),
    }
    return JsonResponse({'ok': True, 'counts': counts})


@login_required
def edit_course(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث بيانات الدورة بنجاح.')
            _log_admin_action(request, f'✏️ تعديل دورة/درس: {course.name}')
            return redirect('admin_panel')
    else:
        form = CourseForm(instance=course)

    context = {
        'form': form,
        'course': course,
        'active_nav': 'admin',
    }
    return render(request, 'halaqat/edit_course.html', context)


@login_required
@require_POST
def delete_course(request, pk):
    course = get_object_or_404(Course, pk=pk)
    name = course.name
    course.delete()
    messages.success(request, f'تم حذف "{name}" وجميع سجلات حضورها.')
    _log_admin_action(request, f'🗑 حذف دورة/درس: {name}')
    return redirect('admin_panel')


# ============================================================
#  صفحة الأساتذة (المرتبطون بحلقة/دورة وغير المرتبطين) وتفقدهم العام
# ============================================================

@login_required
def teachers_list(request):
    """تعرض الأساتذة غير المرتبطين فقط، مع إمكانية تسجيل تفقد عام لمن لم يُؤخذ له حضور اليوم."""
    selected_date = _parse_date_param(request)
    teachers = (
        Teacher.objects.prefetch_related('circles', 'courses')
        .filter(circles__isnull=True, courses__isnull=True)
        .distinct()
        .order_by('name')
    )

    rows = []
    for teacher in teachers:
        general_record = TeacherDailyAttendance.objects.filter(teacher=teacher, date=selected_date).first()
        effective_status = general_record.status if general_record else ''

        rows.append({
            'teacher': teacher,
            'general_status': effective_status,
            'context_notes': [],
            'has_any_record': bool(general_record),
        })

    context = {
        'rows': rows,
        'selected_date': selected_date,
        'total_teachers': teachers.count(),
        'unrecorded_count': sum(1 for r in rows if not r['has_any_record']),
        'active_nav': 'teachers',
    }
    return render(request, 'halaqat/teachers_list.html', context)


@login_required
@require_POST
def save_teacher_attendance_general(request):
    """حفظ فوري لتفقد عام لأستاذ عبر AJAX.

    إن كان الأستاذ مرتبطاً بحلقة أو دورة، نحفظ الحالة في سجلات الحضور المرتبطة بها
    لتُزامَن مع صفحات الحلقة/الدورة. وإلا نحفظ تفقداً عاماً فقط.
    """
    entity_id = request.POST.get('entity_id')
    status = request.POST.get('status')
    date_str = request.POST.get('date')

    if status not in dict(Attendance.STATUS_CHOICES):
        return JsonResponse({'ok': False, 'error': 'حالة غير صالحة'}, status=400)

    try:
        selected_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return JsonResponse({'ok': False, 'error': 'تاريخ غير صالح'}, status=400)

    teacher = get_object_or_404(Teacher, pk=entity_id)
    linked_circles = teacher.circles.all()
    linked_courses = teacher.courses.all()

    if linked_circles.exists() or linked_courses.exists():
        for circle in linked_circles:
            TeacherAttendance.objects.update_or_create(
                teacher=teacher, circle=circle, date=selected_date,
                defaults={'status': status},
            )
        for course in linked_courses:
            CourseTeacherAttendance.objects.update_or_create(
                teacher=teacher, course=course, date=selected_date,
                defaults={'status': status},
            )
    else:
        TeacherDailyAttendance.objects.update_or_create(
            teacher=teacher, date=selected_date,
            defaults={'status': status},
        )

    return JsonResponse({'ok': True})


@login_required
@require_POST
def delete_teacher(request, pk):
    """حذف أستاذ (متاح فقط إن لم يكن مرتبطاً بأي حلقة أو دورة حالياً، لتفادي فقدان البيانات بالخطأ)."""
    teacher = get_object_or_404(Teacher, pk=pk)
    if teacher.circles.exists() or teacher.courses.exists():
        messages.error(request, f'لا يمكن حذف «{teacher.name}» لأنه مرتبط بحلقة أو دورة. عدّل الحلقة/الدورة أولاً.')
    else:
        name = teacher.name
        teacher.delete()
        messages.success(request, f'تم حذف الأستاذ «{name}».')
        _log_admin_action(request, f'🗑 حذف أستاذ: {name}')
    return redirect('teachers_list')


def _queue_whatsapp_absent_messages(selected_date, circle, message_template):
    """إرسال رسائل الغياب بشكل متسلسل في خيط منفصل مع تأخير بعد كل نجاح."""
    records = Attendance.objects.filter(date=selected_date, status='absent')
    if circle:
        records = records.filter(circle=circle)
    records = list(records.select_related('student', 'circle').order_by('circle__name', 'student__name'))

    sent_names = []
    failed_names = []
    skipped_names = []

    for index, record in enumerate(records):
        student = record.student
        phone = (student.phone or '').strip()
        if not phone:
            skipped_names.append(student.name)
            continue

        chat_id = normalize_chat_id(phone)
        if not chat_id:
            failed_names.append(student.name)
            continue

        text = message_template.format(
            student_name=student.name,
            circle_name=record.circle.name,
            date=selected_date.strftime('%Y-%m-%d'),
            week_absences=student.absence_count(7, selected_date),
            month_absences=student.absence_count(30, selected_date),
        )
        if send_whatsapp_text(chat_id, text):
            sent_names.append(student.name)
            if index != len(records) - 1:
                time.sleep(randint(1, 10))
        else:
            failed_names.append(student.name)

    summary_lines = [
        f'📤 تقرير إرسال واتساب للغياب ({selected_date.strftime("%Y-%m-%d")})',
    ]
    if circle:
        summary_lines.append(f'الحلقة: {circle.name}')
    summary_lines.append(f'✅ نجح: {", ".join(sent_names) if sent_names else "لا يوجد"}')
    summary_lines.append(f'❌ فشل/غير مُرسل: {", ".join(failed_names) if failed_names else "لا يوجد"}')
    if skipped_names:
        summary_lines.append(f'⏭ أرقام ناقصة/غير صالحة: {", ".join(skipped_names)}')

    send_telegram_message('\n'.join(summary_lines))


@login_required
def admin_panel(request):
    circle_form = CircleForm()
    student_form = StudentForm()
    course_form = CourseForm()
    whatsapp_form = WhatsAppAbsentForm()

    if request.method == 'POST':
        if 'submit_circle' in request.POST:
            circle_form = CircleForm(request.POST)
            if circle_form.is_valid():
                circle_form.save()
                messages.success(request, 'تمت إضافة الحلقة بنجاح.')
                return redirect('admin_panel')
        elif 'submit_student' in request.POST:
            student_form = StudentForm(request.POST)
            if student_form.is_valid():
                student_form.save()
                messages.success(request, 'تمت إضافة الطالب بنجاح.')
                return redirect('admin_panel')
        elif 'submit_course' in request.POST:
            course_form = CourseForm(request.POST)
            if course_form.is_valid():
                course_form.save()
                messages.success(request, 'تمت إضافة الدورة/الدرس بنجاح.')
                return redirect('admin_panel')
        elif 'submit_whatsapp_absent' in request.POST:
            whatsapp_form = WhatsAppAbsentForm(request.POST)
            if whatsapp_form.is_valid():
                selected_date = whatsapp_form.cleaned_data['date']
                circle = whatsapp_form.cleaned_data['circle']
                message_template = whatsapp_form.cleaned_data['message']

                thread = threading.Thread(
                    target=_queue_whatsapp_absent_messages,
                    args=(selected_date, circle, message_template),
                    daemon=True,
                )
                thread.start()

                messages.success(request, 'تم بدء إرسال الرسائل للغائبين بشكل متسلسل في الخلفية. سيتم إرسال ملخص النتيجة إلى التلغرام عند الانتهاء.')
                return redirect('admin_panel')

    context = {
        'circle_form': circle_form,
        'student_form': student_form,
        'course_form': course_form,
        'whatsapp_form': whatsapp_form,
        'circles': Circle.objects.select_related('teacher').order_by('display_order', 'name', 'id'),
        'students': Student.objects.select_related('circle').all(),
        'courses': Course.objects.prefetch_related('teachers', 'students').all(),
        'active_nav': 'admin',
    }
    return render(request, 'halaqat/admin_panel.html', context)


@login_required
def edit_circle(request, pk):
    circle = get_object_or_404(Circle, pk=pk)
    if request.method == 'POST':
        form = CircleForm(request.POST, instance=circle)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث بيانات الحلقة بنجاح.')
            _log_admin_action(request, f'✏️ تعديل حلقة: {circle.name}')
            return redirect('admin_panel')
    else:
        form = CircleForm(instance=circle)

    context = {
        'form': form,
        'circle': circle,
        'active_nav': 'admin',
    }
    return render(request, 'halaqat/edit_circle.html', context)


@login_required
def edit_student(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث بيانات الطالب بنجاح.')
            _log_admin_action(request, f'✏️ تعديل طالب: {student.name}')
            return redirect('admin_panel')
    else:
        form = StudentForm(instance=student)

    context = {
        'form': form,
        'student': student,
        'active_nav': 'admin',
    }
    return render(request, 'halaqat/edit_student.html', context)


@login_required
@require_POST
def delete_circle(request, pk):
    circle = get_object_or_404(Circle, pk=pk)
    name = circle.name
    circle.delete()  # يحذف تلقائياً طلاب الحلقة وسجلات حضورهم (CASCADE)
    messages.success(request, f'تم حذف حلقة "{name}" وجميع بياناتها المرتبطة بها.')
    _log_admin_action(request, f'🗑 حذف حلقة: {name} (وكل طلابها وسجلات حضورهم)')
    return redirect('admin_panel')


@login_required
@require_POST
def delete_student(request, pk):
    student = get_object_or_404(Student, pk=pk)
    name = student.name
    student.delete()  # يحذف تلقائياً سجلات حضوره (CASCADE)
    messages.success(request, f'تم حذف الطالب "{name}".')
    _log_admin_action(request, f'🗑 حذف طالب: {name}')
    return redirect('admin_panel')


def _parse_date_param(request, param='date'):
    date_str = request.GET.get(param)
    if date_str:
        try:
            return datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    return timezone.now().date()


def _log_admin_action(request, text):
    """يرسل سجل عملية (حذف/تعديل) من لوحة التحكم إلى بوت التلغرام، إن كان مفعّلاً."""
    if not getattr(settings, 'TELEGRAM_LOG_ADMIN_ACTIONS', False):
        return
    who = request.user.get_username() if request.user.is_authenticated else 'غير معروف'
    when = timezone.now().strftime('%Y-%m-%d %H:%M')
    send_telegram_message(f'{text}\nبواسطة: {who}\nالوقت: {when}')


@login_required
def absentees_list(request):
    """عرض الطلاب الغائبين (وقسم للطلاب المسجَّلين إذن) في تاريخ محدد، مع إمكانية تعديل الحالة مباشرة."""
    selected_date = _parse_date_param(request)
    circle_id = request.GET.get('circle')

    def build_groups(status):
        records = Attendance.objects.filter(
            date=selected_date, status=status
        ).select_related('student', 'circle').order_by('circle__name', 'student__name')
        if circle_id:
            records = records.filter(circle_id=circle_id)

        groups = {}
        for record in records:
            groups.setdefault(record.circle, []).append(record.student)

        grouped = []
        for circle, students in groups.items():
            rows = []
            for student in students:
                rows.append({
                    'student': student,
                    'status': status,
                    'week_absences': student.absence_count(7, selected_date),
                    'month_absences': student.absence_count(30, selected_date),
                })
            grouped.append({'circle': circle, 'rows': rows})
        return grouped, records.count()

    grouped_absentees, total_absent = build_groups('absent')
    grouped_excused, total_excused = build_groups('excused')
    grouped_late, total_late = build_groups('late')

    context = {
        'selected_date': selected_date,
        'grouped_absentees': grouped_absentees,
        'grouped_excused': grouped_excused,
        'grouped_late': grouped_late,
        'total_absent': total_absent,
        'total_excused': total_excused,
        'total_late': total_late,
        'circles': Circle.objects.all(),
        'selected_circle_id': int(circle_id) if circle_id else None,
        'active_nav': 'absentees',
    }
    return render(request, 'halaqat/absentees.html', context)


@login_required
def search_students(request):
    """بحث فوري عن طالب بالاسم عبر AJAX، يعيد اسمه واسم حلقته لتحويله إليها مباشرة."""
    query = request.GET.get('q', '').strip()
    results = []
    if query:
        students = (
            Student.objects.select_related('circle')
            .filter(name__icontains=query)
            .order_by('name')[:15]
        )
        results = [
            {
                'id': s.id,
                'name': s.name,
                'circle_id': s.circle_id,
                'circle_name': s.circle.name,
            }
            for s in students
        ]
    return JsonResponse({'results': results})


# ============================================================
#  جرد معلومات الطالب / الأستاذ (بيانات + سجل حضور تفصيلي)
# ============================================================

@login_required
def student_profile(request, student_id):
    student = get_object_or_404(Student.objects.select_related('circle', 'circle__teacher'), pk=student_id)

    circle_records = Attendance.objects.filter(student=student).order_by('-date')
    course_records = CourseAttendance.objects.filter(student=student).select_related('course').order_by('-date')

    stats = {
        'present': circle_records.filter(status='present').count(),
        'late': circle_records.filter(status='late').count(),
        'excused': circle_records.filter(status='excused').count(),
        'absent': circle_records.filter(status='absent').count(),
        'total': circle_records.count(),
    }

    paginator = Paginator(circle_records, 30)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'student': student,
        'stats': stats,
        'rate_30': student.attendance_rate(30),
        'rate_90': student.attendance_rate(90),
        'page_obj': page_obj,
        'course_records': course_records,
        'courses': student.courses.all(),
        'active_nav': None,
    }
    return render(request, 'halaqat/student_profile.html', context)


@login_required
def teacher_profile(request, teacher_id):
    teacher = get_object_or_404(Teacher, pk=teacher_id)

    circle_records = list(TeacherAttendance.objects.filter(teacher=teacher).select_related('circle'))
    course_records = list(CourseTeacherAttendance.objects.filter(teacher=teacher).select_related('course'))
    general_records = list(TeacherDailyAttendance.objects.filter(teacher=teacher))

    combined = []
    for r in circle_records:
        combined.append({'date': r.date, 'status': r.status, 'status_display': r.get_status_display(), 'source': f'حلقة {r.circle.name}'})
    for r in course_records:
        combined.append({'date': r.date, 'status': r.status, 'status_display': r.get_status_display(), 'source': f'دورة {r.course.name}'})
    for r in general_records:
        combined.append({'date': r.date, 'status': r.status, 'status_display': r.get_status_display(), 'source': 'تفقد عام'})

    combined.sort(key=lambda r: r['date'], reverse=True)

    stats = {
        'present': sum(1 for r in combined if r['status'] == 'present'),
        'late': sum(1 for r in combined if r['status'] == 'late'),
        'excused': sum(1 for r in combined if r['status'] == 'excused'),
        'absent': sum(1 for r in combined if r['status'] == 'absent'),
        'total': len(combined),
    }

    paginator = Paginator(combined, 30)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'teacher': teacher,
        'stats': stats,
        'page_obj': page_obj,
        'circles': teacher.circles.all(),
        'courses': teacher.courses.all(),
        'active_nav': None,
    }
    return render(request, 'halaqat/teacher_profile.html', context)


@login_required
def export_attendance_csv(request):
    """تصدير سجل الحضور والغياب ليوم محدد (لكل الحلقات أو حلقة واحدة) كملف CSV."""
    selected_date = _parse_date_param(request)
    circle_id = request.GET.get('circle')

    circles = Circle.objects.select_related('teacher').all()
    if circle_id:
        circles = circles.filter(pk=circle_id)

    buffer = io.StringIO()
    buffer.write('\ufeff')  # BOM ليقرأ Excel النص العربي بشكل صحيح
    writer = csv.writer(buffer)
    writer.writerow(['الحلقة', 'الأستاذ', 'اسم الطالب', 'التاريخ', 'الحالة'])

    for circle in circles:
        if circle.teacher:
            ta = TeacherAttendance.objects.filter(
                teacher=circle.teacher, circle=circle, date=selected_date
            ).first()
            teacher_status_display = ta.get_status_display() if ta else 'لم يُسجَّل'
            writer.writerow([
                circle.name,
                circle.teacher.name,
                f'{circle.teacher.name} (الأستاذ)',
                selected_date.isoformat(),
                teacher_status_display,
            ])

        existing = {
            a.student_id: a.get_status_display()
            for a in Attendance.objects.filter(circle=circle, date=selected_date)
        }
        for student in circle.students.all():
            status_display = existing.get(student.id, 'لم يُسجَّل')
            writer.writerow([
                circle.name,
                circle.teacher.name if circle.teacher else '',
                student.name,
                selected_date.isoformat(),
                status_display,
            ])

    filename = f'attendance_{selected_date.isoformat()}.csv'
    response = HttpResponse(buffer.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_course_attendance_csv(request):
    """تصدير سجل حضور دورة/درس إضافي ليوم محدد كملف CSV."""
    selected_date = _parse_date_param(request)
    course_id = request.GET.get('course')

    courses = Course.objects.prefetch_related('teachers', 'students').all()
    if course_id:
        courses = courses.filter(pk=course_id)

    buffer = io.StringIO()
    buffer.write('\ufeff')
    writer = csv.writer(buffer)
    writer.writerow(['الدورة/الدرس', 'الأستاذ', 'اسم الطالب', 'التاريخ', 'الحالة'])

    for course in courses:
        teacher_names = ', '.join([teacher.name for teacher in course.teachers.all()])
        for teacher in course.teachers.all():
            cta = CourseTeacherAttendance.objects.filter(
                teacher=teacher, course=course, date=selected_date
            ).first()
            teacher_status_display = cta.get_status_display() if cta else 'لم يُسجَّل'
            writer.writerow([
                course.name,
                teacher.name,
                f'{teacher.name} (الأستاذ)',
                selected_date.isoformat(),
                teacher_status_display,
            ])

        existing = {
            a.student_id: a.get_status_display()
            for a in CourseAttendance.objects.filter(course=course, date=selected_date)
        }
        for student in course.students.all():
            status_display = existing.get(student.id, 'لم يُسجَّل')
            writer.writerow([
                course.name,
                teacher_names,
                student.name,
                selected_date.isoformat(),
                status_display,
            ])

    filename = f'course_attendance_{selected_date.isoformat()}.csv'
    response = HttpResponse(buffer.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ============================================================
#  إرسال نسخة احتياطية فورية كاملة إلى تلغرام (Excel)
# ============================================================

@login_required
def send_all_data_telegram(request):
    """
    يُولِّد ملف Excel متعدد الأوراق يحوي جميع بيانات المسجد
    (أساتذة، حلقات، طلاب، حضور، دورات) ويرسله فوراً إلى تلغرام.
    محمي بتسجيل الدخول ويسجّل العملية في لوحة التحكم.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, '❌ مكتبة openpyxl غير مثبتة. شغّل: pip install openpyxl')
        return redirect('admin_panel')

    # ---- ثوابت التنسيق ----
    HEADER_FILL = PatternFill(start_color='B17F4A', end_color='B17F4A', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial')
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _style_sheet(ws, headers, col_widths):
        """يضيف صف العناوين بتنسيق موحّد ويضبط عرض الأعمدة."""
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
        ws.row_dimensions[1].height = 22
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = 'A2'   # تجميد صف العناوين عند التمرير
        ws.sheet_view.rightToLeft = True

    wb = openpyxl.Workbook()
    today = timezone.now().date()
    today_str = today.isoformat()

    # ── ورقة 1: الأساتذة ──────────────────────────────────────
    ws_teachers = wb.active
    ws_teachers.title = 'الأساتذة'
    _style_sheet(ws_teachers, ['م', 'اسم الأستاذ', 'رقم الهاتف'], [5, 30, 20])
    for i, t in enumerate(Teacher.objects.order_by('name'), 1):
        ws_teachers.append([i, t.name, t.phone or '—'])

    # ── ورقة 2: الحلقات ───────────────────────────────────────
    ws_circles = wb.create_sheet('الحلقات')
    _style_sheet(ws_circles,
                 ['م', 'اسم الحلقة', 'الأستاذ', 'عدد الطلاب', 'الوصف', 'ترتيب العرض'],
                 [5, 28, 28, 12, 35, 13])
    for i, c in enumerate(Circle.objects.select_related('teacher').order_by('display_order', 'name'), 1):
        ws_circles.append([i, c.name,
                           c.teacher.name if c.teacher else '—',
                           c.student_count, c.description or '—', c.display_order])

    # ── ورقة 3: الطلاب ────────────────────────────────────────
    ws_students = wb.create_sheet('الطلاب')
    _style_sheet(ws_students,
                 ['م', 'اسم الطالب', 'الحلقة', 'رقم الهاتف', 'ملاحظات', 'تاريخ الانضمام'],
                 [5, 30, 28, 18, 32, 15])
    for i, s in enumerate(
        Student.objects.select_related('circle').order_by('circle__display_order', 'circle__name', 'name'), 1
    ):
        ws_students.append([i, s.name, s.circle.name,
                            s.phone or '—', s.notes or '—',
                            s.joined_at.date().isoformat()])

    # ── ورقة 4: سجل حضور الحلقات ─────────────────────────────
    ws_att = wb.create_sheet('حضور الحلقات')
    _style_sheet(ws_att,
                 ['التاريخ', 'الحلقة', 'النوع', 'الاسم', 'الحالة'],
                 [14, 28, 10, 30, 10])
    for ta in TeacherAttendance.objects.select_related('teacher', 'circle').order_by('-date', 'circle__name'):
        ws_att.append([ta.date.isoformat(), ta.circle.name, 'أستاذ',
                       ta.teacher.name, ta.get_status_display()])
    for a in Attendance.objects.select_related('student', 'circle').order_by('-date', 'circle__name', 'student__name'):
        ws_att.append([a.date.isoformat(), a.circle.name, 'طالب',
                       a.student.name, a.get_status_display()])

    # ── ورقة 5: الدورات ──────────────────────────────────────
    ws_courses = wb.create_sheet('الدورات')
    _style_sheet(ws_courses,
                 ['م', 'اسم الدورة', 'الأساتذة', 'عدد الطلاب', 'الوصف'],
                 [5, 32, 32, 13, 35])
    for i, c in enumerate(Course.objects.prefetch_related('teachers', 'students').order_by('name'), 1):
        teachers_str = '، '.join(t.name for t in c.teachers.all()) or '—'
        ws_courses.append([i, c.name, teachers_str, c.student_count, c.description or '—'])

    # ── ورقة 6: سجل حضور الدورات ─────────────────────────────
    ws_catt = wb.create_sheet('حضور الدورات')
    _style_sheet(ws_catt,
                 ['التاريخ', 'الدورة', 'النوع', 'الاسم', 'الحالة'],
                 [14, 30, 10, 30, 10])
    for cta in CourseTeacherAttendance.objects.select_related('teacher', 'course').order_by('-date', 'course__name'):
        ws_catt.append([cta.date.isoformat(), cta.course.name, 'أستاذ',
                        cta.teacher.name, cta.get_status_display()])
    for ca in CourseAttendance.objects.select_related('student', 'course').order_by('-date', 'course__name', 'student__name'):
        ws_catt.append([ca.date.isoformat(), ca.course.name, 'طالب',
                        ca.student.name, ca.get_status_display()])

    # ── تجميع الملف وإرساله ───────────────────────────────────
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # ملخص في caption التلغرام
    caption = (
        f'📊 بيانات مسجد طه — {today_str}\n'
        f'👥 الطلاب: {Student.objects.count()} | '
        f'📖 الحلقات: {Circle.objects.count()} | '
        f'🎓 الدورات: {Course.objects.count()} | '
        f'👨‍🏫 الأساتذة: {Teacher.objects.count()}'
    )
    export_filename = f'masjid_taha_{today_str}.xlsx'

    ok = send_telegram_document_bytes(excel_buffer.getvalue(), export_filename, caption=caption)

    if ok:
        messages.success(request, '✅ تم إرسال ملف البيانات الكاملة إلى تلغرام بنجاح.')
        _log_admin_action(request, f'📤 إرسال بيانات Excel كاملة إلى تلغرام')
    else:
        messages.error(request, '❌ فشل الإرسال إلى تلغرام. تحقق من إعدادات البوت والاتصال بالإنترنت.')

    return redirect('admin_panel')

# ============================================================
#  ترحيل البيانات من SQLite إلى PostgreSQL
# ============================================================

# ترتيب الحذف والاستيراد يحترم علاقات FK:
#   الحذف:   TeacherDailyAttendance → CourseTeacherAttendance → CourseAttendance
#            → Course (تحذف M2M تلقائياً) → TeacherAttendance → Attendance
#            → Student → Circle → Teacher
#   الاستيراد: عكس الترتيب أعلاه تماماً

_SQLITE_TABLES = [
    # (label_ar, table_name, model_class أو None لجداول M2M)
    ('الأساتذة',            'halaqat_teacher',                    Teacher),
    ('الحلقات',             'halaqat_circle',                     Circle),
    ('الطلاب',              'halaqat_student',                    Student),
    ('حضور الطلاب',         'halaqat_attendance',                 Attendance),
    ('حضور الأساتذة',       'halaqat_teacherattendance',          TeacherAttendance),
    ('الدورات',             'halaqat_course',                     Course),
    ('أساتذة الدورات M2M',  'halaqat_course_teachers',            None),
    ('طلاب الدورات M2M',   'halaqat_course_students',            None),
    ('حضور الدورات',        'halaqat_courseattendance',           CourseAttendance),
    ('حضور أساتذة الدورات', 'halaqat_courseteacherattendance',   CourseTeacherAttendance),
    ('التفقد العام',        'halaqat_teacherdailyattendance',     TeacherDailyAttendance),
]


def _sqlite_count(conn, table):
    """إرجاع عدد الصفوف في جدول SQLite (0 إذا لم يوجد)."""
    try:
        return conn.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
    except Exception:
        return 0


def _sqlite_read(conn, table):
    """
    يقرأ جدول SQLite بأمان ويعيد (مجموعة أسماء الأعمدة المتاحة، قائمة الصفوف).
    يعيد (set(), []) إذا لم يوجد الجدول أو كان فارغاً.
    """
    try:
        cur = conn.execute(f'PRAGMA table_info("{table}")')
        available = {row[1] for row in cur.fetchall()}
        if not available:
            return set(), []
        rows = conn.execute(f'SELECT * FROM "{table}"').fetchall()
        return available, rows
    except Exception:
        return set(), []


def _field_default(model_cls, field_name):
    """
    يعيد القيمة الافتراضية لحقل Django — يُستخدم عند غياب العمود في SQLite القديم.
    يرجع None إذا لم يكن للحقل قيمة افتراضية.
    """
    from django.db import models as dj_models
    try:
        field = model_cls._meta.get_field(field_name)
    except Exception:
        return None
    if field.has_default():
        d = field.default
        return d() if callable(d) else d
    # احتياطي حسب نوع الحقل
    if isinstance(field, (dj_models.CharField, dj_models.TextField)):
        return ''
    if isinstance(field, (dj_models.IntegerField, dj_models.SmallIntegerField,
                          dj_models.PositiveIntegerField, dj_models.PositiveSmallIntegerField)):
        return 0
    if isinstance(field, dj_models.BooleanField):
        return False
    return None


def _col(row, col, available, fallback=None):
    """
    يقرأ قيمة عمود من صف SQLite بأمان مع حالتين:
    1. العمود غائب تماماً (أُضيف لاحقاً في PostgreSQL) → يعيد fallback.
    2. العمود موجود لكن قيمته NULL وتوجد قيمة افتراضية → يعيد fallback.
    هذا يمنع خطأ NOT NULL constraint حين تُرك الحقل فارغاً في السجلات القديمة.
    """
    if col not in available:
        return fallback
    val = row[col]
    # إذا كانت القيمة NULL في SQLite والـ fallback ليس None، استخدم الافتراضي
    if val is None and fallback is not None:
        return fallback
    return val


def _reset_pg_sequences():
    """إعادة ضبط تسلسلات auto-increment في PostgreSQL بعد INSERT بـ IDs محددة يدوياً."""
    from django.db import connection as pg_conn
    if 'postgresql' not in pg_conn.settings_dict.get('ENGINE', ''):
        return  # SQLite في بيئة الاختبار — لا حاجة لضبط تسلسلات
    seq_tables = [
        'halaqat_teacher', 'halaqat_circle', 'halaqat_student',
        'halaqat_attendance', 'halaqat_teacherattendance', 'halaqat_course',
        'halaqat_courseattendance', 'halaqat_courseteacherattendance',
        'halaqat_teacherdailyattendance',
    ]
    with pg_conn.cursor() as cur:
        for table in seq_tables:
            cur.execute(f"""
                SELECT setval(
                    pg_get_serial_sequence('{table}', 'id'),
                    COALESCE((SELECT MAX(id) FROM "{table}"), 1)
                )
            """)


@login_required
def migrate_sqlite_to_postgres(request):
    """
    GET  → معاينة: يُقارن بين عدد السجلات في SQLite و PostgreSQL.
    POST → تنفيذ: يمسح PostgreSQL ثم يستورد كل البيانات من SQLite مع الحفاظ على IDs الأصلية.
    """
    import sqlite3 as sqlite3_lib
    import os

    sqlite_path = os.path.join(settings.BASE_DIR, 'db.sqlite3')

    # ────────────────────────────────────────────────────────
    #  GET: صفحة المعاينة والتأكيد
    # ────────────────────────────────────────────────────────
    if request.method == 'GET':
        file_exists = os.path.exists(sqlite_path)
        preview = []
        total_sqlite = total_pg = 0

        if file_exists:
            conn = sqlite3_lib.connect(sqlite_path)
            pg_counts = {
                'halaqat_teacher':                  Teacher.objects.count(),
                'halaqat_circle':                   Circle.objects.count(),
                'halaqat_student':                  Student.objects.count(),
                'halaqat_attendance':               Attendance.objects.count(),
                'halaqat_teacherattendance':        TeacherAttendance.objects.count(),
                'halaqat_course':                   Course.objects.count(),
                'halaqat_course_teachers':          Course.teachers.through.objects.count(),
                'halaqat_course_students':          Course.students.through.objects.count(),
                'halaqat_courseattendance':         CourseAttendance.objects.count(),
                'halaqat_courseteacherattendance':  CourseTeacherAttendance.objects.count(),
                'halaqat_teacherdailyattendance':   TeacherDailyAttendance.objects.count(),
            }
            for label, table, _ in _SQLITE_TABLES:
                sq = _sqlite_count(conn, table)
                pg = pg_counts.get(table, 0)
                preview.append({'label': label, 'sqlite': sq, 'pg': pg, 'match': sq == pg})
                total_sqlite += sq
                total_pg += pg
            conn.close()

        return render(request, 'halaqat/migrate_sqlite.html', {
            'active_nav': 'admin',
            'sqlite_path': sqlite_path,
            'file_exists': file_exists,
            'preview': preview,
            'total_sqlite': total_sqlite,
            'total_pg': total_pg,
        })

    # ────────────────────────────────────────────────────────
    #  POST: تنفيذ الترحيل
    # ────────────────────────────────────────────────────────
    import sqlite3 as sqlite3_lib

    if not os.path.exists(sqlite_path):
        messages.error(request, f'❌ ملف db.sqlite3 غير موجود في: {sqlite_path}')
        return redirect('migrate_sqlite_to_postgres')

    conn = sqlite3_lib.connect(sqlite_path)
    conn.row_factory = sqlite3_lib.Row
    results = []

    try:
        with transaction.atomic():

            # ── 1. مسح PostgreSQL بالترتيب العكسي لـ FK ──────────
            TeacherDailyAttendance.objects.all().delete()
            CourseTeacherAttendance.objects.all().delete()
            CourseAttendance.objects.all().delete()
            Course.objects.all().delete()        # يحذف M2M تلقائياً
            TeacherAttendance.objects.all().delete()
            Attendance.objects.all().delete()
            Student.objects.all().delete()
            Circle.objects.all().delete()
            Teacher.objects.all().delete()

            # ── 2. Teacher ─────────────────────────────────────────
            avail, rows = _sqlite_read(conn, 'halaqat_teacher')
            Teacher.objects.bulk_create([
                Teacher(
                    id=r['id'],
                    name=r['name'],
                    phone=_col(r, 'phone', avail, _field_default(Teacher, 'phone')),
                ) for r in rows
            ])
            results.append(('الأساتذة', len(rows)))

            # ── 3. Circle ──────────────────────────────────────────
            # مثال: display_order قد يغيب في SQLite القديم → يأخذ قيمته الافتراضية تلقائياً
            avail, rows = _sqlite_read(conn, 'halaqat_circle')
            objs = []
            for r in rows:
                obj = Circle(
                    id=r['id'],
                    name=r['name'],
                    teacher_id=_col(r, 'teacher_id', avail),
                    description=_col(r, 'description', avail, _field_default(Circle, 'description')),
                    display_order=_col(r, 'display_order', avail, _field_default(Circle, 'display_order')),
                )
                obj.created_at = _col(r, 'created_at', avail) or timezone.now().isoformat()
                objs.append(obj)
            Circle.objects.bulk_create(objs)
            results.append(('الحلقات', len(objs)))

            # ── 4. Student ─────────────────────────────────────────
            avail, rows = _sqlite_read(conn, 'halaqat_student')
            objs = []
            for r in rows:
                obj = Student(
                    id=r['id'],
                    name=r['name'],
                    circle_id=r['circle_id'],
                    phone=_col(r, 'phone', avail, _field_default(Student, 'phone')),
                    notes=_col(r, 'notes', avail, _field_default(Student, 'notes')),
                )
                obj.joined_at = _col(r, 'joined_at', avail) or timezone.now().isoformat()
                objs.append(obj)
            Student.objects.bulk_create(objs)
            results.append(('الطلاب', len(objs)))

            # ── 5. Attendance ──────────────────────────────────────
            avail, rows = _sqlite_read(conn, 'halaqat_attendance')
            objs = []
            for r in rows:
                obj = Attendance(
                    id=r['id'],
                    student_id=r['student_id'],
                    circle_id=r['circle_id'],
                    date=r['date'],
                    status=_col(r, 'status', avail, 'present'),
                )
                obj.recorded_at = _col(r, 'recorded_at', avail) or timezone.now().isoformat()
                objs.append(obj)
            Attendance.objects.bulk_create(objs, ignore_conflicts=True)
            results.append(('حضور الطلاب', len(objs)))

            # ── 6. TeacherAttendance ───────────────────────────────
            avail, rows = _sqlite_read(conn, 'halaqat_teacherattendance')
            objs = []
            for r in rows:
                obj = TeacherAttendance(
                    id=r['id'],
                    teacher_id=r['teacher_id'],
                    circle_id=r['circle_id'],
                    date=r['date'],
                    status=_col(r, 'status', avail, 'present'),
                )
                obj.recorded_at = _col(r, 'recorded_at', avail) or timezone.now().isoformat()
                objs.append(obj)
            TeacherAttendance.objects.bulk_create(objs, ignore_conflicts=True)
            results.append(('حضور الأساتذة', len(objs)))

            # ── 7. Course ──────────────────────────────────────────
            avail, rows = _sqlite_read(conn, 'halaqat_course')
            objs = []
            for r in rows:
                obj = Course(
                    id=r['id'],
                    name=r['name'],
                    description=_col(r, 'description', avail, _field_default(Course, 'description')),
                )
                obj.created_at = _col(r, 'created_at', avail) or timezone.now().isoformat()
                objs.append(obj)
            Course.objects.bulk_create(objs)
            results.append(('الدورات', len(objs)))

            # ── 8. Course.teachers M2M ─────────────────────────────
            avail_m, rows_m = _sqlite_read(conn, 'halaqat_course_teachers')
            if rows_m:
                from django.db import connection as pg_conn
                with pg_conn.cursor() as cur:
                    cur.executemany(
                        'INSERT INTO halaqat_course_teachers (course_id, teacher_id) VALUES (%s, %s)',
                        [(r['course_id'], r['teacher_id']) for r in rows_m]
                    )
            results.append(('أساتذة الدورات (M2M)', len(rows_m)))

            # ── 9. Course.students M2M ─────────────────────────────
            avail_m, rows_m = _sqlite_read(conn, 'halaqat_course_students')
            if rows_m:
                from django.db import connection as pg_conn
                with pg_conn.cursor() as cur:
                    cur.executemany(
                        'INSERT INTO halaqat_course_students (course_id, student_id) VALUES (%s, %s)',
                        [(r['course_id'], r['student_id']) for r in rows_m]
                    )
            results.append(('طلاب الدورات (M2M)', len(rows_m)))

            # ── 10. CourseAttendance ───────────────────────────────
            avail, rows = _sqlite_read(conn, 'halaqat_courseattendance')
            objs = []
            for r in rows:
                obj = CourseAttendance(
                    id=r['id'],
                    student_id=r['student_id'],
                    course_id=r['course_id'],
                    date=r['date'],
                    status=_col(r, 'status', avail, 'present'),
                )
                obj.recorded_at = _col(r, 'recorded_at', avail) or timezone.now().isoformat()
                objs.append(obj)
            CourseAttendance.objects.bulk_create(objs, ignore_conflicts=True)
            results.append(('حضور الدورات', len(objs)))

            # ── 11. CourseTeacherAttendance ────────────────────────
            avail, rows = _sqlite_read(conn, 'halaqat_courseteacherattendance')
            objs = []
            for r in rows:
                obj = CourseTeacherAttendance(
                    id=r['id'],
                    teacher_id=r['teacher_id'],
                    course_id=r['course_id'],
                    date=r['date'],
                    status=_col(r, 'status', avail, 'present'),
                )
                obj.recorded_at = _col(r, 'recorded_at', avail) or timezone.now().isoformat()
                objs.append(obj)
            CourseTeacherAttendance.objects.bulk_create(objs, ignore_conflicts=True)
            results.append(('حضور أساتذة الدورات', len(objs)))

            # ── 12. TeacherDailyAttendance ─────────────────────────
            avail, rows = _sqlite_read(conn, 'halaqat_teacherdailyattendance')
            objs = []
            for r in rows:
                obj = TeacherDailyAttendance(
                    id=r['id'],
                    teacher_id=r['teacher_id'],
                    date=r['date'],
                    status=_col(r, 'status', avail, 'present'),
                )
                obj.recorded_at = _col(r, 'recorded_at', avail) or timezone.now().isoformat()
                objs.append(obj)
            TeacherDailyAttendance.objects.bulk_create(objs, ignore_conflicts=True)
            results.append(('التفقد العام', len(objs)))

            # ── 13. إعادة ضبط sequences PostgreSQL ─────────────────
            _reset_pg_sequences()

    except Exception as exc:
        conn.close()
        messages.error(request, f'❌ فشل الترحيل — تم التراجع عن كل التغييرات: {exc}')
        _log_admin_action(request, f'❌ فشل ترحيل SQLite→PostgreSQL: {exc}')
        return redirect('migrate_sqlite_to_postgres')

    conn.close()
    total = sum(c for _, c in results)
    _log_admin_action(request, f'🔄 ترحيل SQLite→PostgreSQL اكتمل ({total} سجل)')

    return render(request, 'halaqat/migrate_sqlite.html', {
        'active_nav': 'admin',
        'sqlite_path': sqlite_path,
        'migration_done': True,
        'results': results,
        'total': total,
    })
